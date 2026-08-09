import io
from datetime import datetime

import pandas as pd
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from database import get_user_currency
from utils.i18n import get_lang, months, op_type_label, result_label, t

# Значения операций в БД (внутренний формат хранения — не локализуются)
_OP_TYPE_TRADE = "Сделка"
_RESULT_WIN = "Win"
_RESULT_LOSS = "Loss"


def generate_excel_bytes(operations, user_id, lang: str = "ru") -> bytes:
    from collections import defaultdict

    # Явно переданный None → язык пользователя; по умолчанию — русский
    lang = lang or get_lang(user_id)

    curr = get_user_currency(user_id)
    sym = {"USD": "$", "EUR": "€", "USDT": "USDT"}.get(curr, "$")
    money_fmt = f'"{sym}"#,##0.00'

    # Локализованные константы — единый источник истины для отображения
    # и для сравнений при форматировании ячеек (работает и для ru, и для en).
    months_dict = months(lang)

    SHEET_SUMMARY = t(lang, "excel.sheet_summary")

    SECTION_OVERALL = t(lang, "excel.section_overall")
    SECTION_MONTHLY = t(lang, "excel.section_monthly")
    SECTION_PAIRS = t(lang, "excel.section_pairs")

    COL_INDICATOR = t(lang, "excel.col_indicator")
    COL_VALUE = t(lang, "excel.col_value")
    COL_MONTH = t(lang, "excel.col_month")
    COL_TRADES = t(lang, "excel.col_trades")
    COL_WINS = t(lang, "excel.col_wins")
    COL_LOSSES = t(lang, "excel.col_losses")
    COL_TOTAL = t(lang, "excel.col_total", currency=curr)
    COL_PAIR = t(lang, "excel.col_pair")

    HEADER_DATE = t(lang, "excel.col_date")
    HEADER_TIME = t(lang, "excel.col_time")
    HEADER_OP_TYPE = t(lang, "excel.col_op_type")
    HEADER_PAIR = t(lang, "excel.col_pair_full")
    HEADER_LOT = t(lang, "excel.col_lot")
    HEADER_SIDE = t(lang, "excel.col_side")
    HEADER_RESULT = t(lang, "excel.col_result")
    HEADER_RISK = t(lang, "excel.col_risk")
    HEADER_COMMISSION = t(lang, "excel.col_commission")
    HEADER_AMOUNT = t(lang, "excel.col_amount", currency=curr)
    HEADER_BALANCE = t(lang, "excel.col_balance", currency=curr)
    HEADER_NOTE = t(lang, "excel.col_note")

    SUMMARY_CURRENT_BALANCE = t(lang, "excel.summary_current_balance")
    SUMMARY_WINRATE = t(lang, "excel.summary_winrate")
    SUMMARY_PROFIT_FACTOR = t(lang, "excel.summary_profit_factor")

    RESULT_WIN_LABEL = t(lang, "result.win")
    RESULT_LOSS_LABEL = t(lang, "result.loss")

    trades = [r for r in operations if r[1] == _OP_TYPE_TRADE]
    total_trades = len(trades)
    wins = sum(1 for r in trades if r[4] == _RESULT_WIN)
    losses = sum(1 for r in trades if r[4] == _RESULT_LOSS)
    winrate = (wins / total_trades) if total_trades > 0 else 0.0
    total_pl = sum(r[5] for r in trades)

    gross_profit = sum(r[5] for r in trades if r[5] > 0)
    gross_loss = abs(sum(r[5] for r in trades if r[5] < 0))
    profit_factor = (
        (gross_profit / gross_loss)
        if gross_loss > 0
        else (gross_profit if gross_profit > 0 else 0.0)
    )

    max_win_streak, max_loss_streak = 0, 0
    curr_win, curr_loss = 0, 0
    for r in trades:
        if r[4] == _RESULT_WIN:
            curr_win += 1
            curr_loss = 0
            if curr_win > max_win_streak:
                max_win_streak = curr_win
        else:
            curr_loss += 1
            curr_win = 0
            if curr_loss > max_loss_streak:
                max_loss_streak = curr_loss

    peak, max_dd = -float('inf'), 0.0
    for r in operations:
        bal = r[6]
        if bal > peak:
            peak = bal
        dd = peak - bal
        if dd > max_dd:
            max_dd = dd

    current_balance = operations[-1][6] if operations else 0.0
    summary_rows = [
        {"Показатель": SUMMARY_CURRENT_BALANCE, "Значение": current_balance},
        {"Показатель": t(lang, "excel.summary_total_trades"), "Значение": total_trades},
        {"Показатель": t(lang, "excel.summary_wins"), "Значение": wins},
        {"Показатель": t(lang, "excel.summary_losses"), "Значение": losses},
        {"Показатель": SUMMARY_WINRATE, "Значение": winrate},
        {
            "Показатель": t(lang, "excel.summary_total_pl", currency=curr),
            "Значение": total_pl,
        },
        {"Показатель": SUMMARY_PROFIT_FACTOR, "Значение": profit_factor},
        {
            "Показатель": t(lang, "excel.summary_max_dd", currency=curr),
            "Значение": max_dd,
        },
        {
            "Показатель": t(lang, "excel.summary_max_win_streak"),
            "Значение": max_win_streak,
        },
        {
            "Показатель": t(lang, "excel.summary_max_loss_streak"),
            "Значение": max_loss_streak,
        },
    ]

    monthly_stats = defaultdict(lambda: {"total": 0, "wins": 0, "losses": 0, "pl": 0.0})
    pair_stats = defaultdict(lambda: {"total": 0, "wins": 0, "losses": 0, "pl": 0.0})
    sheets_data = defaultdict(list)

    for row in operations:
        (
            date_time_str,
            op_type,
            pair,
            lot,
            result,
            amount,
            balance_after,
            note,
            risk_pct,
            side,
            commission,
            _id,
        ) = row
        dt_obj = datetime.strptime(date_time_str, "%Y-%m-%d %H:%M:%S")
        sheet_name = f"{months_dict[dt_obj.month]} {dt_obj.year}"

        sheets_data[sheet_name].append(
            {
                HEADER_DATE: dt_obj.strftime("%d.%m.%Y"),
                HEADER_TIME: dt_obj.strftime("%H:%M:%S"),
                HEADER_OP_TYPE: op_type_label(lang, op_type),
                HEADER_PAIR: pair if pair != "-" else "",
                HEADER_LOT: lot if lot > 0 else "",
                HEADER_SIDE: side if side in ("Buy", "Sell") else "",
                HEADER_RESULT: (
                    result_label(lang, result) if op_type == _OP_TYPE_TRADE else "-"
                ),
                HEADER_RISK: risk_pct if risk_pct > 0 else "",
                HEADER_COMMISSION: commission if commission > 0 else "",
                HEADER_AMOUNT: amount,
                HEADER_BALANCE: balance_after,
                HEADER_NOTE: note if note else "",
            }
        )

        if op_type == _OP_TYPE_TRADE:
            m_key = f"{months_dict[dt_obj.month]} {dt_obj.year}"
            monthly_stats[m_key]["total"] += 1
            pair_stats[pair]["total"] += 1
            if result == _RESULT_WIN:
                monthly_stats[m_key]["wins"] += 1
                pair_stats[pair]["wins"] += 1
            else:
                monthly_stats[m_key]["losses"] += 1
                pair_stats[pair]["losses"] += 1
            monthly_stats[m_key]["pl"] += amount
            pair_stats[pair]["pl"] += amount

    monthly_rows = []
    for m, s in monthly_stats.items():
        wr = (s["wins"] / s["total"]) if s["total"] > 0 else 0.0
        monthly_rows.append(
            {
                "Месяц": m,
                "Сделок": s["total"],
                "Плюсы": s["wins"],
                "Минусы": s["losses"],
                "Винрейт": wr,
                f"Итог ({curr})": s["pl"],
            }
        )

    pair_rows = []
    for p, s in pair_stats.items():
        wr = (s["wins"] / s["total"]) if s["total"] > 0 else 0.0
        pair_rows.append(
            {
                "Пара": p,
                "Сделок": s["total"],
                "Плюсы": s["wins"],
                "Минусы": s["losses"],
                "Винрейт": wr,
                f"Итог ({curr})": s["pl"],
            }
        )

    output = io.BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        ws_summary = writer.book.create_sheet(title=SHEET_SUMMARY, index=0)

        ws_summary.append([SECTION_OVERALL])
        ws_summary.append([COL_INDICATOR, COL_VALUE])
        for item in summary_rows:
            ws_summary.append([item["Показатель"], item["Значение"]])

        ws_summary.append([])
        ws_summary.append([SECTION_MONTHLY])
        ws_summary.append([COL_MONTH, COL_TRADES, COL_WINS, COL_LOSSES, COL_TOTAL])
        for item in monthly_rows:
            ws_summary.append(
                [
                    item["Месяц"],
                    item["Сделок"],
                    item["Плюсы"],
                    item["Минусы"],
                    item["Винрейт"],
                    item[f"Итог ({curr})"],
                ]
            )

        ws_summary.append([])
        ws_summary.append([SECTION_PAIRS])
        ws_summary.append([COL_PAIR, COL_TRADES, COL_WINS, COL_LOSSES, COL_TOTAL])
        for item in pair_rows:
            ws_summary.append(
                [
                    item["Пара"],
                    item["Сделок"],
                    item["Плюсы"],
                    item["Минусы"],
                    item["Винрейт"],
                    item[f"Итог ({curr})"],
                ]
            )

        for row in ws_summary.iter_rows(
            min_row=1, max_row=ws_summary.max_row, min_col=1, max_col=6
        ):
            for cell in row:
                if cell.value in [SECTION_OVERALL, SECTION_MONTHLY, SECTION_PAIRS]:
                    cell.font = Font(bold=True, size=11)
                elif cell.row in [
                    2,
                    len(summary_rows) + 4,
                    len(summary_rows) + len(monthly_rows) + 7,
                ]:
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(
                        start_color="E9ECEF", end_color="E9ECEF", fill_type="solid"
                    )

        for r in range(3, len(summary_rows) + 3):
            val_name = ws_summary.cell(row=r, column=1).value
            val_cell = ws_summary.cell(row=r, column=2)
            if f"({curr})" in str(val_name) or val_name == SUMMARY_CURRENT_BALANCE:
                val_cell.number_format = money_fmt
            elif val_name == SUMMARY_WINRATE:
                val_cell.number_format = '0.0%'
            elif val_name == SUMMARY_PROFIT_FACTOR:
                val_cell.number_format = '0.00'
            else:
                val_cell.number_format = '#,##0'

        for col in ws_summary.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            col_letter = get_column_letter(col[0].column)
            ws_summary.column_dimensions[col_letter].width = max(max_len + 4, 15)

        green_fill = PatternFill(
            start_color="D4EDDA", end_color="D4EDDA", fill_type="solid"
        )
        red_fill = PatternFill(
            start_color="F8D7DA", end_color="F8D7DA", fill_type="solid"
        )

        for sheet_name, rows in sheets_data.items():
            df = pd.DataFrame(rows)
            df.to_excel(writer, index=False, sheet_name=sheet_name)
            worksheet = writer.sheets[sheet_name]
            worksheet.auto_filter.ref = worksheet.dimensions

            for col_idx in range(1, worksheet.max_column + 1):
                col_name = worksheet.cell(row=1, column=col_idx).value
                if col_name in [HEADER_AMOUNT, HEADER_BALANCE, HEADER_COMMISSION]:
                    for r_idx in range(2, worksheet.max_row + 1):
                        worksheet.cell(
                            row=r_idx, column=col_idx
                        ).number_format = money_fmt
                elif col_name == HEADER_RISK:
                    for r_idx in range(2, worksheet.max_row + 1):
                        if worksheet.cell(row=r_idx, column=col_idx).value:
                            worksheet.cell(
                                row=r_idx, column=col_idx
                            ).number_format = '0.0"%"'

            for row_idx in range(2, worksheet.max_row + 1):
                val = None
                for c_idx in range(1, worksheet.max_column + 1):
                    if worksheet.cell(row=1, column=c_idx).value == HEADER_RESULT:
                        val = worksheet.cell(row=row_idx, column=c_idx).value
                        break
                if val == RESULT_WIN_LABEL:
                    for c in range(1, worksheet.max_column + 1):
                        worksheet.cell(row=row_idx, column=c).fill = green_fill
                elif val == RESULT_LOSS_LABEL:
                    for c in range(1, worksheet.max_column + 1):
                        worksheet.cell(row=row_idx, column=c).fill = red_fill

            for col in worksheet.columns:
                max_len = max(len(str(cell.value or '')) for cell in col)
                worksheet.column_dimensions[
                    get_column_letter(col[0].column)
                ].width = max(max_len + 4, 12)

    return output.getvalue()
